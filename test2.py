# This script takes in raw csv files and generates structured invoices
import pandas as pd
import glob
import openpyxl
from fpdf import FPDF
from pathlib import Path

filepaths = glob.glob("invoices/*.xlsx")
print(filepaths)

# Loop through the filepaths
for filepath in filepaths:
    # Add a page for each iteration
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()

    # Pull the invoice number out of the filename
    filename = Path(filepath).stem
    invoice_nr, date = filename.split("-")

    # Establish font for each sheet
    pdf.set_font(family="Times", size=15, style="B")
    pdf.cell(w=50, h=8, txt=f"Invoices nr.{invoice_nr}", ln=1)

    pdf.set_font(family="Times", size=15, style="B")
    pdf.cell(w=50, h=8, txt=f"Date: {date}", ln=1)

    # Add header
    df = pd.read_excel(filepath, sheet_name="Sheet 1")
    pdf.set_font(family="Times", size=10, style="B")
    columns = [column.replace("_", " ").title() for column in list(df.columns)]
    pdf.cell(w=30, h=8, txt=columns[0], border=1)
    pdf.cell(w=70, h=8, txt=columns[1], border=1)
    pdf.cell(w=35, h=8, txt=columns[2], border=1)
    pdf.cell(w=30, h=8, txt=columns[3], border=1)
    pdf.cell(w=30, h=8, txt=columns[4], border=1, ln=1)
    # Add rows
    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, txt=str(row["product_id"]), border=1) # type: ignore
        pdf.cell(w=70, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=35, h=8, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["total_price"]), border=1, ln=1)

    # noinspection PyArgumentList
    pdf.set_font(family="Times", size=10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=70, h=8, txt="", border=1)
    pdf.cell(w=35, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt=str(df["total_price"].sum()), ln=1)

    pdf.set_font(family="Times", size=10)


    pdf.output(f"PDFs/{filename}.pdf")